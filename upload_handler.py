import warnings
from openpyxl import load_workbook
from io import BytesIO

# Ignore these specific openpyxl warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.worksheet._reader")

def parse_excel(file_bytes):
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = row[1]  # 'NAME' column
        phone = row[6]  # 'Contact number' column

        if name is None or phone is None:
            continue

        try:
            phone_str = str(int(float(phone)))
        except (ValueError, TypeError):
            phone_str = str(phone).strip()

        rows.append((str(name).strip(), phone_str))

    return rows
